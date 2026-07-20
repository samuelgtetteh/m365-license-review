"""Excel report writer (openpyxl).

Sheets:
  1. Summary                 — tenant info, how-to-read, optimization headline
  2. License Optimization    — prioritized recommendations + $ savings (the deliverable)
  3. License Inventory       — purchased / assigned / slack / cost per SKU
  4. Detail: Disabled Users  — R1
  5. Detail: Inactive Users  — R2
  6. Raw Data                — hidden; all users + licenses for auditability

Styling is deliberately restrained: bold headers, frozen header row, USD
formatting, red highlight on high-severity rows. Client-readable, not flashy.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from m365_review.core.models import AuditResult, Severity
from m365_review.settings import get_settings

# --- palette ---
_NAVY = "1F3864"
_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16, color=_NAVY)
_SUBTLE = Font(color="595959")
_HIGH_FILL = PatternFill("solid", fgColor="F8CBAD")      # soft red
_MED_FILL = PatternFill("solid", fgColor="FFE699")       # soft amber
_MONEY = "$#,##0.00"
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_SEV_FILL = {Severity.high: _HIGH_FILL, Severity.medium: _MED_FILL}


def write_xlsx(result: AuditResult, path: Path) -> Path:
    wb = Workbook()
    _summary_sheet(wb.active, result)
    _optimization_sheet(wb.create_sheet("License Optimization"), result)
    _inventory_sheet(wb.create_sheet("License Inventory"), result)
    _expirations_sheet(wb.create_sheet("Subscription Expirations"), result)
    _disabled_detail_sheet(wb.create_sheet("Detail - Disabled Users"), result)
    _inactive_detail_sheet(wb.create_sheet("Detail - Inactive Users"), result)
    _mfa_sheet(wb.create_sheet("Detail - MFA"), result)
    _admin_roles_sheet(wb.create_sheet("Detail - Admin Roles"), result)
    _raw_sheet(wb.create_sheet("Raw Data"), result)
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #

def _header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------------- #
# 1 - Summary
# --------------------------------------------------------------------------- #

def _summary_sheet(ws: Worksheet, result: AuditResult) -> None:
    ws.title = "Summary"
    ws["A1"] = "Microsoft 365 License Optimization Report"
    ws["A1"].font = _TITLE_FONT

    ws["A3"] = "Tenant"
    ws["B3"] = result.tenant_display_name
    ws["A4"] = "Tenant ID"
    ws["B4"] = result.tenant_id
    ws["A5"] = "Report date"
    ws["B5"] = result.report_date.isoformat()
    for r in range(3, 6):
        ws.cell(row=r, column=1).font = Font(bold=True)

    company = get_settings().report_company_name
    if company:
        ws["A2"] = f"Prepared by {company}"
        ws["A2"].font = _SUBTLE

    ws["A7"] = "How to read this report"
    ws["A7"].font = Font(bold=True, size=12, color=_NAVY)
    intro = (
        "This report reviews the Microsoft 365 licenses your tenant owns against how they are "
        "actually used, and lists specific, prioritized actions to reduce cost. Savings figures "
        "are estimates based on the price list in effect at generation time; confirm each "
        "recommendation before acting. The tool is read-only and made no changes to your tenant."
    )
    ws["A8"] = intro
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A8:F11")

    # Optimization headline
    ws["A13"] = "License optimization at a glance"
    ws["A13"].font = Font(bold=True, size=12, color=_NAVY)
    rows = [
        ("Estimated monthly savings identified", result.total_monthly_savings_usd, _MONEY),
        ("Estimated annual savings identified", result.total_annual_savings_usd, _MONEY),
        ("Total licenses purchased", result.total_purchased, None),
        ("Total licenses assigned", result.total_assigned, None),
        ("Total findings", len(result.findings), None),
    ]
    r = 14
    for label, value, fmt in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=value)
        if fmt:
            c.number_format = fmt
        r += 1

    # Severity breakdown
    r += 1
    ws.cell(row=r, column=1, value="Findings by severity").font = Font(bold=True, size=12, color=_NAVY)
    r += 1
    sev_first = r
    for sev, count in result.severity_counts().items():
        ws.cell(row=r, column=1, value=sev.capitalize())
        ws.cell(row=r, column=2, value=count)
        r += 1
    sev_last = r - 1

    # Native pie chart of findings-by-severity (anchored to the right, cols A-F free).
    if result.findings:
        chart = PieChart()
        chart.title = "Findings by severity"
        chart.height = 6.5
        chart.width = 10
        labels = Reference(ws, min_col=1, min_row=sev_first, max_row=sev_last)
        data = Reference(ws, min_col=2, min_row=sev_first, max_row=sev_last)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        ws.add_chart(chart, "H2")

    if result.caveats:
        r += 1
        ws.cell(row=r, column=1, value="Notes & data caveats").font = Font(bold=True, size=12, color=_NAVY)
        r += 1
        for cav in result.caveats:
            cell = ws.cell(row=r, column=1, value=f"• {cav}")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = _SUBTLE
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1

    _autosize(ws, [40, 30, 14, 14, 14, 14])


# --------------------------------------------------------------------------- #
# 2 - License Optimization (the deliverable)
# --------------------------------------------------------------------------- #

def _optimization_sheet(ws: Worksheet, result: AuditResult) -> None:
    headers = [
        "Priority", "Rule", "Severity", "Recommendation / finding",
        "Affected users", "Est. monthly savings", "Est. annual savings",
    ]
    _header_row(ws, 1, headers)
    row = 2
    for i, f in enumerate(result.findings, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=f.rule_id)
        ws.cell(row=row, column=3, value=f.severity.value.capitalize())
        rec = ws.cell(row=row, column=4, value=f"{f.title} — {f.recommendation}")
        rec.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=5, value=f.affected_count)
        m = ws.cell(row=row, column=6, value=f.estimated_monthly_savings_usd)
        m.number_format = _MONEY
        a = ws.cell(row=row, column=7, value=f.estimated_annual_savings_usd)
        a.number_format = _MONEY
        fill = _SEV_FILL.get(f.severity)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        row += 1

    # Total row
    ws.cell(row=row, column=4, value="TOTAL").font = Font(bold=True)
    tot_m = ws.cell(row=row, column=6, value=result.total_monthly_savings_usd)
    tot_m.font = Font(bold=True)
    tot_m.number_format = _MONEY
    tot_a = ws.cell(row=row, column=7, value=result.total_annual_savings_usd)
    tot_a.font = Font(bold=True)
    tot_a.number_format = _MONEY

    _autosize(ws, [8, 8, 10, 70, 14, 18, 18])


# --------------------------------------------------------------------------- #
# 3 - License Inventory
# --------------------------------------------------------------------------- #

def _inventory_sheet(ws: Worksheet, result: AuditResult) -> None:
    headers = [
        "Product", "SKU part number", "Purchased", "Assigned", "Unused",
        "Unit price (mo)", "Monthly cost", "Annual cost", "Unused cost (mo)",
    ]
    _header_row(ws, 1, headers)
    n = len(headers)

    paid = [r for r in result.inventory if not r.is_unlimited]
    free = [r for r in result.inventory if r.is_unlimited]

    row = _inventory_section(ws, 2, n, "PAID LICENSES", paid, show_costs=True)
    if free:
        row += 1  # spacer
        row = _inventory_section(
            ws, row, n, "FREE / SELF-SERVICE LICENSES (excluded from totals)", free, show_costs=False
        )
    _autosize(ws, [34, 30, 11, 10, 9, 15, 14, 14, 15])


def _inventory_section(ws: Worksheet, row: int, ncols: int, title: str, rows, *, show_costs: bool) -> int:
    """Write a labeled category banner + its rows. Returns the next free row."""
    banner_fill = PatternFill("solid", fgColor="404040")
    banner = ws.cell(row=row, column=1, value=title)
    banner.font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = banner_fill
    row += 1

    if not rows:
        ws.cell(row=row, column=1, value="(none)").font = _SUBTLE
        return row + 1

    monthly_subtotal = 0.0
    for r in rows:
        ws.cell(row=row, column=1, value=r.display_name)
        ws.cell(row=row, column=2, value=r.sku_part_number)
        ws.cell(row=row, column=3, value=r.purchased)
        ws.cell(row=row, column=4, value=r.assigned)
        ws.cell(row=row, column=5, value=r.slack)
        if show_costs:
            up = ws.cell(row=row, column=6, value=r.unit_price_usd if r.price_known else "?")
            mc = ws.cell(row=row, column=7, value=r.monthly_cost_usd)
            ac = ws.cell(
                row=row, column=8,
                value=round(r.monthly_cost_usd * 12, 2) if r.monthly_cost_usd is not None else None,
            )
            sc = ws.cell(row=row, column=9, value=r.slack_monthly_usd)
            for c in (up, mc, ac, sc):
                if isinstance(c.value, (int, float)):
                    c.number_format = _MONEY
            monthly_subtotal += r.monthly_cost_usd or 0.0
        else:
            # Free SKUs: no meaningful price/cost.
            for col in (6, 7, 8, 9):
                cell = ws.cell(row=row, column=col, value="Free" if col == 6 else "—")
                cell.font = _SUBTLE
        row += 1

    if show_costs:
        ws.cell(row=row, column=1, value="Paid subtotal (monthly)").font = Font(bold=True)
        sub = ws.cell(row=row, column=7, value=round(monthly_subtotal, 2))
        sub.font = Font(bold=True)
        sub.number_format = _MONEY
        row += 1
    return row


# --------------------------------------------------------------------------- #
# Subscription expirations (its own section)
# --------------------------------------------------------------------------- #

def _expirations_sheet(ws: Worksheet, result: AuditResult) -> None:
    headers = [
        "Product", "SKU part number", "Status", "Licenses", "Trial",
        "Created", "Renews / expires", "Days remaining",
    ]
    _header_row(ws, 1, headers)

    if not result.subscriptions_available:
        ws.cell(row=2, column=1, value="Subscription/expiration data was unavailable for this tenant.").font = _SUBTLE
        _autosize(ws, [34, 26, 12, 10, 8, 14, 18, 14])
        return
    if not result.subscriptions:
        ws.cell(row=2, column=1, value="(no subscriptions returned)").font = _SUBTLE
        _autosize(ws, [34, 26, 12, 10, 8, 14, 18, 14])
        return

    now = result.generated_at
    row = 2
    for s in result.subscriptions:
        days = s.days_until_expiry(now)
        ws.cell(row=row, column=1, value=s.display_name or s.sku_part_number)
        ws.cell(row=row, column=2, value=s.sku_part_number)
        ws.cell(row=row, column=3, value=s.status or "")
        ws.cell(row=row, column=4, value=s.total_licenses)
        ws.cell(row=row, column=5, value="Yes" if s.is_trial else "")
        ws.cell(row=row, column=6, value=s.created_datetime.date().isoformat() if s.created_datetime else "")
        ws.cell(
            row=row, column=7,
            value=s.next_lifecycle_datetime.date().isoformat() if s.next_lifecycle_datetime else "",
        )
        ws.cell(row=row, column=8, value=days if days is not None else "")
        # colour: red if expired, amber if within 30 days
        fill = None
        if days is not None and days < 0:
            fill = _HIGH_FILL
        elif days is not None and days <= 30:
            fill = _MED_FILL
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        row += 1
    _autosize(ws, [34, 26, 12, 10, 8, 14, 18, 14])


# --------------------------------------------------------------------------- #
# 4 / 5 - Detail sheets
# --------------------------------------------------------------------------- #

def _detail_from_rule(ws: Worksheet, result: AuditResult, rule_id: str, third_col: str) -> None:
    headers = ["Display name", "User principal name", third_col, "Monthly cost"]
    _header_row(ws, 1, headers)
    row = 2
    for f in result.findings:
        if f.rule_id != rule_id:
            continue
        for u in f.affected_users:
            ws.cell(row=row, column=1, value=u.display_name)
            ws.cell(row=row, column=2, value=u.user_principal_name)
            ws.cell(row=row, column=3, value=u.detail)
            c = ws.cell(row=row, column=4, value=u.monthly_cost_usd)
            if isinstance(c.value, (int, float)):
                c.number_format = _MONEY
            row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="(no findings for this rule)").font = _SUBTLE
    _autosize(ws, [28, 42, 42, 14])


def _disabled_detail_sheet(ws: Worksheet, result: AuditResult) -> None:
    _detail_from_rule(ws, result, "R1", "Licenses assigned")


def _inactive_detail_sheet(ws: Worksheet, result: AuditResult) -> None:
    _detail_from_rule(ws, result, "R2", "Activity")


# --------------------------------------------------------------------------- #
# Identity security detail (MFA + admin roles)
# --------------------------------------------------------------------------- #

def _mfa_sheet(ws: Worksheet, result: AuditResult) -> None:
    headers = ["Display name", "User principal name", "Type", "Admin", "MFA registered",
               "Default method", "Methods registered", "SSPR reg."]
    widths = [26, 40, 9, 7, 14, 18, 30, 10]
    _header_row(ws, 1, headers)
    data = result.tenant_data
    reg = data.user_registration if data else []
    if data and not data.mfa_data_available:
        ws.cell(row=2, column=1, value="MFA registration data was unavailable for this tenant.").font = _SUBTLE
        _autosize(ws, widths); return
    if not reg:
        ws.cell(row=2, column=1, value="(no registration data)").font = _SUBTLE
        _autosize(ws, widths); return
    row = 2
    for u in sorted(reg, key=lambda r: (r.is_mfa_registered, not r.is_admin)):
        ws.cell(row=row, column=1, value=u.display_name)
        ws.cell(row=row, column=2, value=u.user_principal_name)
        ws.cell(row=row, column=3, value=(u.user_type or "").capitalize())
        ws.cell(row=row, column=4, value="Yes" if u.is_admin else "")
        ws.cell(row=row, column=5, value="Yes" if u.is_mfa_registered else "No")
        ws.cell(row=row, column=6, value=u.default_mfa_method or "")
        ws.cell(row=row, column=7, value=", ".join(u.methods_registered))
        ws.cell(row=row, column=8, value="Yes" if u.is_sspr_registered else "")
        if not u.is_mfa_registered:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = _HIGH_FILL if u.is_admin else _MED_FILL
        row += 1
    _autosize(ws, widths)


def _admin_roles_sheet(ws: Worksheet, result: AuditResult) -> None:
    headers = ["Role", "Member", "User principal name"]
    _header_row(ws, 1, headers)
    data = result.tenant_data
    roles = [r for r in (data.directory_roles if data else []) if r.is_privileged]
    if data and not data.roles_available:
        ws.cell(row=2, column=1, value="Directory role data was unavailable for this tenant.").font = _SUBTLE
        _autosize(ws, [34, 28, 42]); return
    if not roles:
        ws.cell(row=2, column=1, value="(no privileged roles returned)").font = _SUBTLE
        _autosize(ws, [34, 28, 42]); return
    row = 2
    for role in sorted(roles, key=lambda r: r.display_name):
        if not role.members:
            ws.cell(row=row, column=1, value=role.display_name)
            ws.cell(row=row, column=2, value="(no members)").font = _SUBTLE
            row += 1
            continue
        for m in role.members:
            ws.cell(row=row, column=1, value=role.display_name)
            ws.cell(row=row, column=2, value=m.display_name)
            ws.cell(row=row, column=3, value=m.user_principal_name)
            row += 1
    _autosize(ws, [34, 28, 42])


# --------------------------------------------------------------------------- #
# 6 - Raw data (hidden)
# --------------------------------------------------------------------------- #

def _raw_sheet(ws: Worksheet, result: AuditResult) -> None:
    headers = [
        "Display name", "UPN", "Enabled", "Type", "Created",
        "Last sign-in", "Licenses",
    ]
    _header_row(ws, 1, headers)
    row = 2
    data = result.tenant_data
    if data:
        name_map = result.sku_id_to_name
        for u in data.users:
            ws.cell(row=row, column=1, value=u.display_name)
            ws.cell(row=row, column=2, value=u.user_principal_name)
            ws.cell(row=row, column=3, value="Yes" if u.account_enabled else "No")
            ws.cell(row=row, column=4, value=u.user_type)
            ws.cell(row=row, column=5, value=u.created_datetime.isoformat() if u.created_datetime else "")
            last = u.effective_last_sign_in
            ws.cell(row=row, column=6, value=last.isoformat() if last else "")
            # Friendly license names rather than raw GUIDs.
            names = [name_map.get(sid, sid) for sid in u.assigned_license_sku_ids]
            ws.cell(row=row, column=7, value=", ".join(names))
            row += 1
    ws.sheet_state = "hidden"
    _autosize(ws, [26, 40, 9, 10, 22, 22, 46])
