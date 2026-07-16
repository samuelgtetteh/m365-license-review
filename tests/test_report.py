"""Smoke tests for the report writers.

These open the generated files and assert structure (sheet names, row counts,
JSON keys), not pixel-level formatting.
"""

from __future__ import annotations

import json

from m365_review.core.engine import build_result
from m365_review.core.report import write_reports
from m365_review.core.report.json_writer import build_json_payload


def _result(tenant_data, pricing, now):
    return build_result(tenant_data, pricing=pricing, now=now)


def test_writes_all_three_formats(tenant_data, pricing, now, tmp_path):
    result = _result(tenant_data, pricing, now)
    paths = write_reports(result, output_dir=tmp_path, formats=["xlsx", "docx", "json"])
    assert set(paths) == {"xlsx", "docx", "json"}
    for p in paths.values():
        assert p.exists() and p.stat().st_size > 0


def test_xlsx_sheets_and_hidden_raw(tenant_data, pricing, now, tmp_path):
    from openpyxl import load_workbook

    result = _result(tenant_data, pricing, now)
    paths = write_reports(result, output_dir=tmp_path, formats=["xlsx"])
    wb = load_workbook(paths["xlsx"])
    assert wb.sheetnames == [
        "Summary",
        "License Optimization",
        "License Inventory",
        "Detail - Disabled Users",
        "Detail - Inactive Users",
        "Raw Data",
    ]
    assert wb["Raw Data"].sheet_state == "hidden"


def test_json_optimization_summary(tenant_data, pricing, now):
    result = _result(tenant_data, pricing, now)
    payload = build_json_payload(result)
    summ = payload["optimization_summary"]
    assert summ["total_estimated_monthly_savings_usd"] == 729.0
    assert summ["total_estimated_annual_savings_usd"] == 8748.0
    assert summ["finding_count"] == len(payload["findings"])
    assert payload["tenant"]["display_name"] == "Contoso Ltd"


def test_json_roundtrips(tenant_data, pricing, now, tmp_path):
    result = _result(tenant_data, pricing, now)
    paths = write_reports(result, output_dir=tmp_path, formats=["json"])
    data = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert data["license_inventory"]
    assert "caveats" in data


def test_base_filename_is_safe(tenant_data, pricing, now):
    result = _result(tenant_data, pricing, now)
    assert result.base_filename() == "Contoso_Ltd_2026-07-16"
